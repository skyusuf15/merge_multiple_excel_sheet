# Program to merge multiple excel sheets to one

def decorator(f):
    def wrapper():
        print('::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::')
        print('Starting Program Process.....')
        f()
        print('Process Completed.....')
        print('::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::')

    return wrapper


@decorator
def merge_sheets():

    from openpyxl import load_workbook
    from pandas import DataFrame
    import glob

    # fetch all files in input folder 
    filesInDirectory = glob.glob('./input/*.xlsx')
    # iterate over the files
    for filepath in filesInDirectory:

        # Get the file name
        fileName = filepath.split('/')[-1]

        # Load excel sheet into memory
        workbook = load_workbook(filepath)

        # Initialize variable list data to store all sheet rows
        list_data = []

        # Iterate over all sheets
        for sheet in workbook.sheetnames:
            # convert the current sheet to list and remove first 5 rows (the header)
            toBeProcessed = list(workbook[sheet].values)[4:]

            # # check if a sheet has a record
            # if toBeProcessed[0][0] is not None:

            # store sheet's row to list_data variable
            list_data.extend(toBeProcessed)

        # filter out empty cells
        list_data = list(filter(lambda row: row[0] is not None, list_data))

        # convert list_data to dataframe and convert to excel
        DataFrame(list_data).to_excel('./output/' + fileName)


merge_sheets()
